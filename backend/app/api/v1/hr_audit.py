from typing import List, Optional
from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.database.session import get_db
from app.models.user import User
from app.models.audit_log import AuditLog
from app.security.rbac import get_current_hr
from app.schemas.hr_operations import AuditLogResponse

router = APIRouter()

@router.get("/", response_model=List[AuditLogResponse])
def get_audit_logs(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 100,
    action: Optional[str] = None,
    role: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_hr)
):
    query = db.query(AuditLog)
    
    if action:
        query = query.filter(AuditLog.action == action)
    if role:
        query = query.filter(AuditLog.role == role)
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                AuditLog.action.ilike(search_filter),
                AuditLog.details.ilike(search_filter)
            )
        )
        
    logs = query.order_by(AuditLog.timestamp.desc()).offset(skip).limit(limit).all()
    
    result = []
    for log in logs:
        result.append(AuditLogResponse(
            id=log.id,
            user_id=log.user_id,
            action=log.action,
            role=log.role or "Unknown",
            details=log.details or "",
            timestamp=log.timestamp
        ))
        
    return result

import csv
import io
from fastapi.responses import StreamingResponse

from fastapi import Path

@router.get("/export/{format}")
def export_audit_logs(
    format: str = Path(..., description="Format can be csv, excel, or pdf"),
    db: Session = Depends(get_db),
    action: Optional[str] = None,
    role: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(get_current_hr)
):
    query = db.query(AuditLog)
    
    if action:
        query = query.filter(AuditLog.action == action)
    if role:
        query = query.filter(AuditLog.role == role)
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                AuditLog.action.ilike(search_filter),
                AuditLog.details.ilike(search_filter)
            )
        )
        
    logs = query.order_by(AuditLog.timestamp.desc()).all()
    
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "User ID", "Action", "Role", "Details", "Timestamp"])
        
        for log in logs:
            writer.writerow([
                log.id, 
                log.user_id, 
                log.action, 
                log.role, 
                log.details, 
                log.timestamp.isoformat() if log.timestamp else ""
            ])
            
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=audit_logs.csv"}
        )
        
    elif format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font
        import tempfile
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"
        
        headers = ["ID", "User ID", "Action", "Role", "Details", "Timestamp"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        for log in logs:
            ws.append([
                log.id, 
                log.user_id, 
                log.action, 
                log.role, 
                log.details, 
                log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else ""
            ])
            
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp_name = tmp.name
            
        def iterfile():
            with open(tmp_name, "rb") as f:
                yield from f
                
        return StreamingResponse(
            iterfile(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=audit_logs.xlsx"}
        )
        
    elif format == "pdf":
        from fpdf import FPDF
        import tempfile
        
        pdf = FPDF(orientation='L') # Landscape for more width
        pdf.add_page()
        pdf.set_font("Arial", size=14, style='B')
        pdf.cell(270, 10, txt="Enterprise HRMS - Audit Logs", ln=True, align='C')
        pdf.ln(5)
        
        # Header
        pdf.set_font("Arial", size=9, style='B')
        pdf.cell(15, 10, txt="ID", border=1)
        pdf.cell(15, 10, txt="User", border=1)
        pdf.cell(45, 10, txt="Action", border=1)
        pdf.cell(25, 10, txt="Role", border=1)
        pdf.cell(125, 10, txt="Details", border=1)
        pdf.cell(45, 10, txt="Timestamp", border=1, ln=True)
        
        # Data
        pdf.set_font("Arial", size=8)
        for log in logs:
            pdf.cell(15, 8, txt=str(log.id), border=1)
            pdf.cell(15, 8, txt=str(log.user_id), border=1)
            pdf.cell(45, 8, txt=str(log.action)[:30], border=1)
            pdf.cell(25, 8, txt=str(log.role), border=1)
            pdf.cell(125, 8, txt=str(log.details)[:90], border=1)
            pdf.cell(45, 8, txt=log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "", border=1, ln=True)
            
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf.output(tmp.name)
            tmp_name = tmp.name
            
        def iterfile():
            with open(tmp_name, "rb") as f:
                yield from f
                
        return StreamingResponse(
            iterfile(),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=audit_logs.pdf"}
        )
    else:
        return {"error": "Invalid format requested"}
