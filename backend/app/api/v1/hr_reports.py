from datetime import datetime, timezone
from typing import List
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
import io
import csv
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database.session import get_db
from app.models.user import User
from app.models.department import Department
from app.models.leave_balance import LeaveBalance
from app.models.leave_request import LeaveRequest
from app.models.leave_type import LeaveType
from app.models.employee import Employee
from app.security.rbac import get_current_hr
from app.schemas.hr_operations import DepartmentLeaveReport, LeaveTypeReport

router = APIRouter()

@router.get("/leave/departments", response_model=List[DepartmentLeaveReport])
def report_leave_by_department(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_hr)
):
    current_year = datetime.now(timezone.utc).year
    departments = db.query(Department).all()
    
    result = []
    for dept in departments:
        # Get all employees in dept
        emp_ids = [e.id for e in dept.employees]
        if not emp_ids:
            continue
            
        balances = db.query(LeaveBalance).filter(
            LeaveBalance.employee_id.in_(emp_ids),
            LeaveBalance.calendar_year == current_year
        ).all()
        
        total_eligible = sum([b.eligible for b in balances])
        total_used = sum([b.used for b in balances])
        total_available = sum([b.available for b in balances])
        total_pending = sum([b.pending for b in balances])
        
        result.append(DepartmentLeaveReport(
            department_name=dept.department_name,
            total_eligible=total_eligible,
            total_used=total_used,
            total_available=total_available,
            total_pending=total_pending
        ))
        
    return result

@router.get("/leave/types", response_model=List[LeaveTypeReport])
def report_leave_by_type(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_hr)
):
    current_year = datetime.now(timezone.utc).year
    leave_types = db.query(LeaveType).all()
    
    result = []
    for lt in leave_types:
        requests = db.query(LeaveRequest).filter(
            LeaveRequest.leave_type_id == lt.id,
            func.extract('year', LeaveRequest.applied_on) == current_year
        ).all()
        
        total_reqs = len(requests)
        approved = len([r for r in requests if r.status == "Approved"])
        rejected = len([r for r in requests if r.status == "Rejected"])
        
        result.append(LeaveTypeReport(
            leave_type_name=lt.leave_name,
            total_requests=total_reqs,
            approved_requests=approved,
            rejected_requests=rejected
        ))
        
    return result

@router.get("/dashboard/export")
def export_dashboard_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_hr)
):
    current_year = datetime.now(timezone.utc).year
    
    # Gather metrics
    total_employees = db.query(func.count(Employee.id)).scalar() or 0
    active_employees = db.query(func.count(Employee.id)).filter(Employee.employment_status == "Active").scalar() or 0
    on_leave = db.query(func.count(Employee.id)).filter(Employee.employment_status == "On Leave").scalar() or 0
    pending_approvals = db.query(func.count(LeaveRequest.id)).filter(LeaveRequest.status == "Awaiting HR").scalar() or 0
    
    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(["Enterprise HRMS - Dashboard Report"])
    writer.writerow([f"Generated On: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([])
    
    writer.writerow(["Metric", "Value"])
    writer.writerow(["Total Employees", total_employees])
    writer.writerow(["Active Employees", active_employees])
    writer.writerow(["Employees On Leave", on_leave])
    writer.writerow(["Pending HR Approvals", pending_approvals])
    writer.writerow([])
    
    # Department summary
    writer.writerow(["Department", "Total Eligible Leave", "Total Used Leave", "Total Available Leave", "Total Pending", "Total LWP Days Used"])
    departments = db.query(Department).all()
    for dept in departments:
        emp_ids = [e.id for e in dept.employees]
        if not emp_ids:
            continue
        balances = db.query(LeaveBalance).filter(
            LeaveBalance.employee_id.in_(emp_ids),
            LeaveBalance.calendar_year == current_year
        ).all()
        lwp_used = db.query(func.sum(LeaveRequest.lwp_days)).filter(
            LeaveRequest.employee_id.in_(emp_ids),
            LeaveRequest.status == "Approved",
            func.extract('year', LeaveRequest.start_date) == current_year
        ).scalar() or 0.0
        
        writer.writerow([
            dept.department_name,
            sum([b.eligible for b in balances]),
            sum([b.used for b in balances]),
            sum([b.available for b in balances]),
            sum([b.pending for b in balances]),
            lwp_used
        ])
    
    output.seek(0)
    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=hr_dashboard_report.csv"
    return response

from fastapi import Path

@router.get("/dashboard/export/{format}")
def export_advanced_report(
    format: str = Path(..., description="Format can be csv, excel, or pdf"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_hr)
):
    current_year = datetime.now(timezone.utc).year
    
    # Gather metrics
    total_employees = db.query(func.count(Employee.id)).scalar() or 0
    active_employees = db.query(func.count(Employee.id)).filter(Employee.employment_status == "Active").scalar() or 0
    on_leave = db.query(func.count(Employee.id)).filter(Employee.employment_status == "On Leave").scalar() or 0
    pending_approvals = db.query(func.count(LeaveRequest.id)).filter(LeaveRequest.status == "Awaiting HR").scalar() or 0
    
    # Department summary
    departments = db.query(Department).all()
    dept_data = []
    for dept in departments:
        emp_ids = [e.id for e in dept.employees]
        if not emp_ids:
            continue
        balances = db.query(LeaveBalance).filter(
            LeaveBalance.employee_id.in_(emp_ids),
            LeaveBalance.calendar_year == current_year
        ).all()
        lwp_used = db.query(func.sum(LeaveRequest.lwp_days)).filter(
            LeaveRequest.employee_id.in_(emp_ids),
            LeaveRequest.status == "Approved",
            func.extract('year', LeaveRequest.start_date) == current_year
        ).scalar() or 0.0
        dept_data.append({
            "name": dept.department_name,
            "eligible": sum([b.eligible for b in balances]),
            "used": sum([b.used for b in balances]),
            "available": sum([b.available for b in balances]),
            "pending": sum([b.pending for b in balances]),
            "lwp_used": lwp_used
        })
        
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Enterprise HRMS - Detailed Report"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Employees", total_employees])
        writer.writerow(["Active Employees", active_employees])
        writer.writerow(["Employees On Leave", on_leave])
        writer.writerow(["Pending HR Approvals", pending_approvals])
        writer.writerow([])
        writer.writerow(["Department", "Eligible", "Used", "Available", "Pending", "LWP Used"])
        for d in dept_data:
            writer.writerow([d["name"], d["eligible"], d["used"], d["available"], d["pending"], d["lwp_used"]])
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]), 
            media_type="text/csv", 
            headers={"Content-Disposition": "attachment; filename=hr_report.csv"}
        )
        
    elif format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import tempfile
        
        wb = Workbook()
        ws = wb.active
        ws.title = "HR Report"
        
        ws.append(["Enterprise HRMS - Detailed Report"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        
        ws.append(["Metric", "Value"])
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)
        ws.append(["Total Employees", total_employees])
        ws.append(["Active Employees", active_employees])
        ws.append(["Employees On Leave", on_leave])
        ws.append(["Pending HR Approvals", pending_approvals])
        ws.append([])
        
        ws.append(["Department", "Eligible", "Used", "Available", "Pending", "LWP Used"])
        for cell in ws[9]:
            cell.font = Font(bold=True)
            
        for d in dept_data:
            ws.append([d["name"], d["eligible"], d["used"], d["available"], d["pending"], d["lwp_used"]])
            
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp_name = tmp.name
            
        def iterfile():
            with open(tmp_name, "rb") as f:
                yield from f
                
        return StreamingResponse(
            iterfile(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=hr_report.xlsx"}
        )
        
    elif format == "pdf":
        from fpdf import FPDF
        import tempfile
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=16, style='B')
        pdf.cell(200, 10, txt="Enterprise HRMS - Detailed Report", ln=True, align='C')
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, txt=f"Generated On: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
        pdf.ln(10)
        
        pdf.set_font("Arial", size=12, style='B')
        pdf.cell(100, 10, txt="Summary Metrics", ln=True)
        pdf.set_font("Arial", size=11)
        pdf.cell(100, 10, txt=f"Total Employees: {total_employees}", ln=True)
        pdf.cell(100, 10, txt=f"Active Employees: {active_employees}", ln=True)
        pdf.cell(100, 10, txt=f"Employees On Leave: {on_leave}", ln=True)
        pdf.cell(100, 10, txt=f"Pending HR Approvals: {pending_approvals}", ln=True)
        pdf.ln(10)
        
        pdf.set_font("Arial", size=12, style='B')
        pdf.cell(100, 10, txt="Department Breakdown", ln=True)
        pdf.set_font("Arial", size=10, style='B')
        pdf.cell(40, 10, txt="Department", border=1)
        pdf.cell(30, 10, txt="Eligible", border=1)
        pdf.cell(30, 10, txt="Used", border=1)
        pdf.cell(30, 10, txt="Available", border=1)
        pdf.cell(30, 10, txt="Pending", border=1)
        pdf.cell(30, 10, txt="LWP", border=1, ln=True)
        
        pdf.set_font("Arial", size=10)
        for d in dept_data:
            pdf.cell(40, 10, txt=str(d["name"]), border=1)
            pdf.cell(30, 10, txt=str(d["eligible"]), border=1)
            pdf.cell(30, 10, txt=str(d["used"]), border=1)
            pdf.cell(30, 10, txt=str(d["available"]), border=1)
            pdf.cell(30, 10, txt=str(d["pending"]), border=1)
            pdf.cell(30, 10, txt=str(d["lwp_used"]), border=1, ln=True)
            
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf.output(tmp.name)
            tmp_name = tmp.name
            
        def iterfile():
            with open(tmp_name, "rb") as f:
                yield from f
                
        return StreamingResponse(
            iterfile(),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=hr_report.pdf"}
        )
        
    else:
        return {"error": "Invalid format requested"}
